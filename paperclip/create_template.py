"""
Generate a sample Excel template showing the expected column format.

Run:  python -m paperclip.create_template
"""
import sys
from pathlib import Path


def create_template(output_path: str = "birth_data_template.xlsx") -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl required: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Birth Data"

    headers = [
        "name", "date", "time", "city", "country",
        "latitude", "longitude", "timezone", "notes"
    ]

    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample data rows
    samples = [
        ["Albert Einstein", "1879-03-14", "11:30:00", "Ulm", "Germany", 48.3974, 9.9934, "+1.0", "Famous physicist"],
        ["Mahatma Gandhi", "1869-10-02", "07:45:00", "Porbandar", "India", 21.6417, 69.6293, "+5.5", ""],
        ["Marie Curie", "1867-11-07", "12:00:00", "Warsaw", "Poland", 52.2297, 21.0122, "+1.0", ""],
        ["Your Name Here", "YYYY-MM-DD", "HH:MM:SS", "City", "Country", 0.0, 0.0, "+0.0", ""],
    ]

    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Set column widths
    widths = [20, 14, 10, 15, 12, 12, 12, 10, 25]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Notes sheet
    notes_ws = wb.create_sheet("Format Notes")
    notes = [
        ("Column", "Format", "Examples"),
        ("name", "Free text", "Albert Einstein"),
        ("date", "YYYY-MM-DD or DD/MM/YYYY", "1879-03-14 or 14/03/1879"),
        ("time", "HH:MM or HH:MM:SS", "11:30 or 11:30:00"),
        ("city", "City name", "London"),
        ("country", "Country name", "United Kingdom"),
        ("latitude", "Decimal degrees (N=+, S=-)", "51.5074 or -33.8688"),
        ("longitude", "Decimal degrees (E=+, W=-)", "-0.1278 or 151.2093"),
        ("timezone", "UTC offset or name", "+5.5 or IST or America/New_York"),
        ("notes", "Optional free text", "Any notes about this chart"),
    ]
    for row in notes:
        notes_ws.append(row)

    wb.save(output_path)
    print(f"Template saved to: {output_path}")


if __name__ == "__main__":
    create_template()
